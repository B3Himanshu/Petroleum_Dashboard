"""
HSRL Dashboard — Accounting Data Map Generator
==============================================

Generates a master Excel governance schema mapping internal accounting GL/nominal
codes to the HSRL Petroleum Dashboard's live UI components, charts, KPIs and
API payload keys. Output is intended for executive / client review.

Output file:
    HSRL_Dashboard_Accounting_Data_Map_Master.xlsx
    (written next to this script)

Usage:
    pip install pandas openpyxl
    python Documentation/generate_data_map.py

Styling:
    - Deep Navy header (#1F3864) with white bold text
    - Center alignment for Column 1 (GL code)
    - Wrapped text for the long-form Logic column
    - Monospace style for the API payload-key column
    - Auto-fit column widths (computed from max content per column)
    - Thin gridlines on every cell + frozen header row
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Output configuration
# ----------------------------------------------------------------------------
OUTPUT_PATH = Path(__file__).parent / "HSRL_Dashboard_Accounting_Data_Map_Master.xlsx"
SHEET_NAME = "Data Governance Map"

COLUMNS = [
    "Internal Accounting GL/Nominal Code",
    "Company Account Ledger Name",
    "HSRL Dashboard Module / Pillar",
    "Live UI Component / Chart Widget Name",
    "Data Extraction & Functional Transformation Logic",
    "Production API Frontend Payload Key",
]

# ----------------------------------------------------------------------------
# Master data rows
# (gl_code, ledger_name, module, component, logic_in_plain_english, payload_key)
# ----------------------------------------------------------------------------
ROWS = [
    # ============== FUEL — sales (£) and volume (litres) ==============
    ("4000", "Unleaded Petrol", "Dashboard · Fuel Pillar",
     "Net Sales KPI Card · Fuel Grade Mix Chart · Date-Wise Sales Chart",
     "Total Unleaded petrol sales for the selected date range. Litres sold are read from the transaction notes attached to each entry.",
     "fuelSales · fuelVolume · netSalesBreakdown[].4000"),

    ("4001", "Diesel", "Dashboard · Fuel Pillar",
     "Net Sales KPI · Fuel Grade Mix · Bunkered Sales Chart",
     "Total Diesel sales for the period. Litres are read from the transaction notes. Bunkered Diesel litres come from code 4101 (volume only, not added to sales £).",
     "fuelSales · fuelVolume"),

    ("4002", "Super Unleaded", "Dashboard · Fuel Pillar",
     "Net Sales KPI · Fuel Grade Mix Chart",
     "Total Super Unleaded sales for the period. Litres are read from the transaction notes.",
     "fuelSales · fuelVolume"),

    ("4003", "Super Diesel", "Dashboard · Fuel Pillar",
     "Net Sales KPI · Fuel Grade Mix Chart",
     "Total Super Diesel sales for the period. Litres are read from the transaction notes.",
     "fuelSales · fuelVolume"),

    ("4004", "Adblue", "Dashboard · Fuel Pillar",
     "Net Sales KPI · Fuel Grade Mix Chart",
     "Total Adblue sales for the period. Litres are read from the transaction notes.",
     "fuelSales · fuelVolume"),

    # ---------------- Bunkering ----------------
    ("4100", "Bunkering Charges - BP Commission", "Dashboard · Fuel Pillar",
     "Bunkered Sales Chart · Total Site Revenue KPI",
     "BP bunkering commission income. Counted in Total Site Revenue but excluded from headline Fuel Sales £.",
     "bunkeredBreakdown[].4100 · totalSiteRevenue"),

    ("4101", "Bunkered Sales", "Dashboard · Fuel Pillar",
     "Fuel Volume KPI · Bunkered Sales Chart",
     "Volume-only line for bunkered fuel. Reversal entries (Rev.Accrual) are automatically subtracted. Not added to fuel sales £.",
     "fuelVolume"),

    ("4102", "Bunkered Commission", "Dashboard · Fuel Pillar",
     "Bunkered Sales Chart",
     "Commission income from bunkered fuel transactions.",
     "bunkeredBreakdown[].4102"),

    # ---------------- Fuel purchase / cost ----------------
    ("5000", "Unleaded - Purchase", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · Net Profit Card · EBITDA Card",
     "Cost of Unleaded petrol purchased from suppliers. Subtracted from Unleaded sales to calculate fuel margin.",
     "fuelCost · fuelProfit"),

    ("5001", "Diesel - Purchase", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · Net Profit Card",
     "Cost of Diesel purchased. Paired with Diesel sales (4001) to calculate diesel margin.",
     "fuelCost · fuelProfit"),

    ("5002", "Super Unleaded - Purchase", "Dashboard · Fuel Pillar",
     "Fuel Profit Card",
     "Cost of Super Unleaded purchased. Paired with sales code 4002.",
     "fuelCost · fuelProfit"),

    ("5003", "Super Diesel - Purchase", "Dashboard · Fuel Pillar",
     "Fuel Profit Card",
     "Cost of Super Diesel purchased. Paired with sales code 4003.",
     "fuelCost · fuelProfit"),

    ("5004", "Adblue - Purchase", "Dashboard · Fuel Pillar",
     "Fuel Profit Card",
     "Cost of Adblue purchased. Paired with sales code 4004.",
     "fuelCost · fuelProfit"),

    ("5005", "Fuel Promotional", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · Net Profit Card",
     "Promotional discount cost on fuel sales (e.g. loyalty programmes, pump price drops). Reduces fuel profit.",
     "fuelCost · fuelProfit"),

    ("5041", "Fuel Commission", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · Net Profit Card",
     "Fuel-related commission paid out. Reduces fuel profit.",
     "fuelCost · fuelProfit"),

    ("5046", "Fuel Stock Movement - Unleaded (inferred)", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · ROI Card",
     "Stock movement / adjustment for Unleaded. Please confirm exact name with your chart of accounts.",
     "fuelCost · fuelProfit"),

    ("5047", "Fuel Stock Movement - Diesel (inferred)", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · ROI Card",
     "Stock movement / adjustment for Diesel. Please confirm exact name with your chart of accounts.",
     "fuelCost · fuelProfit"),

    ("5048", "Fuel Stock Movement - Super Unleaded (inferred)", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · ROI Card",
     "Stock movement / adjustment for Super Unleaded. Please confirm exact name with your chart of accounts.",
     "fuelCost · fuelProfit"),

    ("5049", "Fuel Stock Movement - Super Diesel (inferred)", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · ROI Card",
     "Stock movement / adjustment for Super Diesel. Please confirm exact name with your chart of accounts.",
     "fuelCost · fuelProfit"),

    ("5050", "Stock Movement (general)", "Dashboard · Fuel Pillar",
     "Fuel Profit Card · ROI Card",
     "General stock movement / adjustment line for fuel.",
     "fuelCost · fuelProfit"),

    # ============== SHOP — sales ==============
    ("4032", "E-Pay Sales", "Dashboard · Shop Pillar",
     "Shop Profit Card · GP Breakdown Modal",
     "E-Pay sales income (mobile top-ups, utility payments).",
     "shopSales · salesBreakdown[].4032"),

    ("4034", "Paypoint / Keycharge Sales", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Paypoint / Keycharge sales income.",
     "shopSales"),

    ("4036", "Lottery Online", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Online lottery sales income.",
     "shopSales"),

    ("4037", "Lottery Instants", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Instant lottery (scratchcards) sales income.",
     "shopSales"),

    ("4039", "EV Revenue", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "EV charging revenue. Paired with cost code 5039.",
     "shopSales"),

    # ============== SHOP — cost ==============
    ("5016", "Grocery - Purchase", "Dashboard · Shop Pillar",
     "Shop Profit Card · GP Breakdown Modal",
     "Grocery purchase cost. Added to Shop Costs in May 2026 - was previously excluded. Reduces Shop Profit.",
     "shopCost · costBreakdown[].5016"),

    ("5032", "E-Pay Purchases", "Dashboard · Shop Pillar",
     "Shop Profit Card · GP Breakdown Modal",
     "E-Pay purchase cost. Paired with sales code 4032.",
     "shopCost · costBreakdown[].5032"),

    ("5033", "E-Pay Commission", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Commission paid on E-Pay transactions.",
     "shopCost"),

    ("5034", "Paypoint / Keycharge Purchases", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Paypoint / Keycharge purchase cost. Paired with sales code 4034.",
     "shopCost"),

    ("5035", "Paypoint / Keycharge Commissions", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Commission paid on Paypoint / Keycharge transactions.",
     "shopCost"),

    ("5036", "Lottery Online - Cost", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Cost of online lottery sales. Paired with sales code 4036.",
     "shopCost"),

    ("5037", "Lottery Instants - Cost", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Cost of instant lottery (scratchcards). Paired with sales code 4037.",
     "shopCost"),

    ("5039", "EV Costs", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "EV charging costs (electricity, network fees). Paired with sales code 4039.",
     "shopCost"),

    ("5042", "Lottery (Instant + Online) Commission", "Dashboard · Shop Pillar",
     "Shop Profit Card",
     "Combined lottery commission payments to the National Lottery operator.",
     "shopCost"),

    # ============== VALET / COFFEE — sales ==============
    ("4017", "Hot Food / Costa - Sales", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card · GP Breakdown Modal",
     "Hot food and Costa coffee sales income. Paired with cost code 5015.",
     "valetSales · salesBreakdown[].4017"),

    ("4028", "Car Wash", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Car Wash sales income. Paired with cost code 5028.",
     "valetSales"),

    ("4029", "Jet Wash", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Jet Wash sales income. Paired with cost code 5029.",
     "valetSales"),

    ("4030", "Car Vac", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Car vacuum sales income. Paired with cost code 5030.",
     "valetSales"),

    ("4031", "Car Airline", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Car tyre air pump income. Paired with cost code 5031.",
     "valetSales"),

    # ============== VALET / COFFEE — cost ==============
    ("5015", "Hot Food / Costa - Purchase", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Cost of hot food and Costa coffee stock. Paired with sales code 4017.",
     "valetCost"),

    ("5028", "Car Wash - Cost", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Car Wash running costs (chemicals, water, maintenance).",
     "valetCost"),

    ("5029", "Jet Wash - Cost", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Jet Wash running costs.",
     "valetCost"),

    ("5030", "Car Vac - Cost", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Car Vac running costs (electricity, maintenance).",
     "valetCost"),

    ("5031", "Car Airline - Cost", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Car Airline (tyre pump) running costs.",
     "valetCost"),

    ("5043", "Valet Commission", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Commission paid on valet services to third parties.",
     "valetCost"),

    ("5044", "Coffee Commission", "Dashboard · Coffee & Valet Pillar",
     "Valet Profit Card",
     "Commission paid on Costa coffee sales to the brand operator.",
     "valetCost"),

    # ============== MISC INCOME — 13 codes feeding EBITDA ==============
    ("4400", "Marketing Services Income", "Dashboard · EBITDA Pillar",
     "EBITDA Card · EBITDA Breakdown Modal (Misc Income line)",
     "Income from marketing services. Counts towards Other (Misc) Income in EBITDA.",
     "miscIncome"),

    ("4401", "ATM Cash Machine Income Received", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Income received from on-site ATM operators.",
     "miscIncome"),

    ("4402", "Rebates - Income", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Supplier rebates received during the period.",
     "miscIncome"),

    ("4404", "Commissions Received", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Commission income from third-party services.",
     "miscIncome"),

    ("4405", "Insurance Claims & Compensations", "Dashboard · EBITDA Pillar",
     "EBITDA Card · EBITDA Breakdown Modal",
     "Insurance payouts and compensation received. Added to Misc Income in May 2026 - was previously excluded.",
     "miscIncome"),

    ("4407", "Rental Income", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Rent received from property and forecourt sublets. One of the largest steady Misc Income lines (around £89-98k/month combined).",
     "miscIncome"),

    ("4410", "Misc Income - ALL HEAD OFFICE", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Other head office miscellaneous income. Sometimes contains one-off bookings (e.g. £44k spike in Jan 2026).",
     "miscIncome"),

    ("4412", "Ast-Costa Coffee Rent", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Rent received from Costa Coffee at Astwick site.",
     "miscIncome"),

    ("4413", "EV Rent / Revenue", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Rent / revenue share received from EV charge-point operators.",
     "miscIncome"),

    ("4415", "Bank Interest Income", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Interest earned on bank deposits. The single largest Misc Income line (around £90-94k/month).",
     "miscIncome"),

    ("4416", "ByBox Income", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Income from ByBox parcel locker installations.",
     "miscIncome"),

    ("4417", "Amazon Locker Rent", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Rent received from Amazon Locker installations.",
     "miscIncome"),

    ("4418", "Euro Car Parks Rebate", "Dashboard · EBITDA Pillar",
     "EBITDA Card",
     "Rebate income from Euro Car Parks for forecourt parking enforcement.",
     "miscIncome"),

    # ============== LABOUR ==============
    ("7000", "Gross Wages", "Dashboard · Labour Pillar",
     "Labour Cost KPI Card · Labour % chart",
     "Gross wages paid to all site and support staff. Used in Labour Cost % = Labour / Fuel Sales x 100.",
     "totalLabourCost · labourCostPercent"),

    ("7001", "Employer's NI - Staff", "Dashboard · Labour Pillar",
     "Labour Cost KPI Card",
     "Employer's National Insurance contributions on staff wages.",
     "totalLabourCost"),

    ("7002", "Directors' Salaries", "Dashboard · Labour Pillar",
     "Labour Cost KPI Card",
     "Salaries paid to company directors.",
     "totalLabourCost"),

    ("7003", "Employer's NI - Directors", "Dashboard · Labour Pillar",
     "Labour Cost KPI Card",
     "Employer's NI on directors' salaries.",
     "totalLabourCost"),

    ("7005", "Directors' Pensions", "Dashboard · Labour Pillar",
     "Labour Cost KPI Card",
     "Pension contributions on behalf of directors.",
     "totalLabourCost"),

    # ============== EBITA OVERHEADS ==============
    ("7150", "Rent", "Dashboard · EBITDA Pillar",
     "EBITDA Card · Actual PPL Card · Overhead Cost Breakdown Modal",
     "Property rent paid. One of the five operational overheads subtracted from EBITDA.",
     "totalOverheads · overheadsBreakdown[].7150"),

    ("7151", "Rates", "Dashboard · EBITDA Pillar",
     "EBITDA Card · Actual PPL Card · Overhead Cost Breakdown Modal",
     "Business rates paid to councils. Operational overhead in EBITDA.",
     "totalOverheads · overheadsBreakdown[].7151"),

    ("7200", "Light & Heat", "Dashboard · EBITDA Pillar",
     "EBITDA Card · Actual PPL Card · Overhead Cost Breakdown Modal",
     "Electricity, gas and heating costs across all sites. Operational overhead in EBITDA.",
     "totalOverheads · overheadsBreakdown[].7200"),

    ("7800", "Repairs & Renewals", "Dashboard · EBITDA Pillar",
     "EBITDA Card · Actual PPL Card · Overhead Cost Breakdown Modal",
     "Property repairs and equipment renewals. Operational overhead in EBITDA.",
     "totalOverheads · overheadsBreakdown[].7800"),

    ("7906", "Credit Card Charges", "Dashboard · EBITDA Pillar",
     "EBITDA Card · Actual PPL Card · Overhead Cost Breakdown Modal",
     "Credit / debit card processing fees. Operational overhead in EBITDA.",
     "totalOverheads · overheadsBreakdown[].7906"),

    # ============== DEPRECIATION ==============
    ("8200", "Depreciation - Motor Vehicles", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal (Depreciation section)",
     "Depreciation expense on motor vehicles. Subtracted from EBITDA to derive Total Net Profit.",
     "depreciation · depreciationBreakdown[].8200"),

    ("8201", "Depreciation - Leasehold L&B", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Depreciation expense on leasehold land and buildings. Subtracted from EBITDA.",
     "depreciation · depreciationBreakdown[].8201"),

    ("8202", "Depreciation - Freehold Land & Building", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Depreciation expense on freehold land and buildings. Subtracted from EBITDA.",
     "depreciation · depreciationBreakdown[].8202"),

    ("8203", "Depreciation - Plant & Machinery", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Depreciation expense on plant and machinery (pumps, tanks). Subtracted from EBITDA.",
     "depreciation · depreciationBreakdown[].8203"),

    ("8204", "Depreciation - Fixtures & Fittings", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Depreciation expense on fixtures and fittings. Subtracted from EBITDA.",
     "depreciation · depreciationBreakdown[].8204"),

    ("8206", "Depreciation - Other Assets", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Depreciation expense on other assets. Subtracted from EBITDA. (Note: code 8205 is not used in this chart of accounts.)",
     "depreciation · depreciationBreakdown[].8206"),

    ("8207", "Depreciation - Site Development & Improvement", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Depreciation expense on site development and improvement works. Subtracted from EBITDA.",
     "depreciation · depreciationBreakdown[].8207"),

    # ============== LOAN INTEREST + CORP TAX ==============
    ("7750", "Loan Interest Paid", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Interest paid on loans. Subtracted from EBITDA. Kept separate from Overheads so PPL after Overheads is not double-counted.",
     "loanInterest"),

    ("9000", "Corporation Tax Charge", "Dashboard · Net Profit Pillar",
     "Total Net Profit Breakdown Modal",
     "Corporation tax charge. Final deduction in: Total Net Profit = EBITDA - Depreciation - Loan Interest - Corporation Tax.",
     "corporationTax"),

    # ============== KPI CARDS (aggregated metrics) ==============
    ("4000-4005, 5000-5005, 5041, 5046-5050, 4032/4034/4036/4037/4039, 5016, 5032-5037, 5039, 5042, 4017, 4028-4031, 5015, 5028-5031, 5043, 5044",
     "Aggregate: Fuel Profit + Shop Profit (signed) + Valet Profit (signed)",
     "Dashboard · Top KPI Strip",
     "Gross Profit KPI Card",
     "Total Gross Profit across the business = Fuel Profit + Shop Profit + Coffee & Valet Profit. Shop or Valet losses correctly reduce the total (May 2026 fix).",
     "grossProfit · fuelProfit · shopProfit · valetProfit"),

    ("Revenue: 4000-4004, 4100-4102  |  Cost: 5000-5005, 5041, 5046-5050",
     "Aggregate: Net Profit = Total Revenue - Total Cost",
     "Dashboard · Top KPI Strip",
     "Net Profit KPI Card",
     "Net Profit = Total Revenue minus Total Cost across the 14-code Net Profit set. Sage stores costs as negative numbers, so the code adds the raw values which arithmetically yields Revenue minus Cost.",
     "totalProfit · netProfit · totalRevenue · totalCost"),

    ("(All Gross Profit codes) + 4400, 4401, 4402, 4404, 4405, 4407, 4410, 4412, 4413, 4415, 4416, 4417, 4418 - 7150, 7151, 7200, 7800, 7906",
     "Aggregate: Gross Profit + Misc Income (13 codes incl. 4405) - EBITA Overheads",
     "Dashboard · Quick Insights",
     "EBITDA KPI Card",
     "Earnings before interest, tax, depreciation and amortisation. Calculation: Gross Profit + Other (Misc) Income - Operating Overheads. Misc Income widened to 13 codes (added 4405) in May 2026.",
     "ebita"),

    ("4000-4004 (litres parsed from details) + 4101 (bunker volume) + 5000-5005, 5041, 5046-5050 (cost)",
     "Derived: Fuel Profit / Fuel Volume x 100 (pence per litre)",
     "Dashboard · Quick Insights",
     "Gross PPL KPI Card (Avg PPL)",
     "Profit made per litre of fuel sold, shown in pence. Calculation: Fuel Profit divided by Litres Sold, multiplied by 100. If volume is missing, the calculation falls back to sales value.",
     "avgPPL"),

    ("4000-4004, 5000-5004 (Fuel Profit) + 7150, 7151, 7200, 7800, 7906 (EBITA Overheads)",
     "Derived: (Fuel Profit - Overheads) / Volume x 100",
     "Dashboard · Quick Insights",
     "PPL after O/H KPI Card (Actual PPL)",
     "Profit per litre AFTER deducting operating overheads (rent, rates, light & heat, repairs, card fees). Shows the true margin per litre after running costs.",
     "pplAfterOverheads · actualPPL"),

    ("(All EBITDA codes) - 8200-8207 (Depreciation) - 7750 (Loan Interest) - 9000 (Corp Tax)  |  Investment N/Cs configured in code",
     "Derived: Total Net Profit / Investment x 100",
     "Dashboard · ROI",
     "ROI KPI Card · ROI Monthly Trend Chart",
     "Return on Investment as a percentage. Calculation: Total Net Profit divided by Investment, shown as a percent.",
     "roi · roiMonthlyTrend"),

    ("dept_number DISTINCT  |  Revenue evidence: 4000-4004",
     "Aggregate: COUNT(DISTINCT dept_number) WHERE revenue rows exist",
     "Dashboard · Quick Insights",
     "Active Sites KPI Card",
     "Count of petrol stations that recorded any fuel sales in the selected period. Head Office is excluded from the count.",
     "activeSites"),

    ("1200, 1201, 1202, 1203, 1204, 1211-1222, 1240, 1250, 1251",
     "Aggregate: Latest balance across HO LLOYDS + per-site bank accounts (1200-series)",
     "Dashboard · Quick Insights",
     "Bank Balance KPI Card",
     "Most recent bank balance across all company accounts (head office Lloyds account, individual site accounts, deposit accounts).",
     "totalBankBalance · bankBalanceBreakdown[]"),

    ("4000-4004 (totalNetSales) / DISTINCT dept_number",
     "Derived: Total Net Sales / Active Sites",
     "Dashboard · Quick Insights",
     "Average Sale Per Site KPI Card",
     "Average sales value per active station = Total Net Sales divided by the number of active sites in the period.",
     "avgSalePerSite"),

    # ============== CHARTS ==============
    ("4000-4004, 5000-5005, 5041, 5046-5050, 4032/4034/4036/4037/4039, 5016, 5032-5037, 5039, 5042, 4017, 4028-4031, 5015, 5028-5031, 5043, 5044, 4400-4418, 7150, 7151, 7200, 7800, 7906  (grouped by month)",
     "Aggregate Chart: Monthly GP + EBITDA + Volume time-series",
     "Dashboard · Trends",
     "Monthly Performance Chart",
     "Trend chart showing Gross Profit, EBITDA and Fuel Volume month by month across the selected window.",
     "charts.monthlyPerformance[]"),

    ("4000-4004, 5000-5005 (grouped by sage_date day)",
     "Aggregate Chart: Daily Sales + Volume + Profit",
     "Dashboard · Trends",
     "Date-Wise Data Chart",
     "Daily breakdown of sales, fuel volume and profit across the selected date range.",
     "charts.dateWise[]"),

    ("Unleaded: 4000+5000+5046  |  Diesel: 4001+4101+5001+5041+5047+4100+4102+5005  |  Super Unleaded: 4002+5002+5048  |  Super Diesel: 4003+5003+5049  |  Adblue: 4004+5004+5050",
     "Per-Grade Margin Mix: Unleaded / Diesel / Super Unleaded / Super Diesel / Adblue / Shop / Coffee & Valet",
     "Dashboard · Fuel Pillar",
     "Fuel Grade Mix Chart (Pie)",
     "Pie chart showing each fuel grade's share of total fuel profit, plus Shop and Coffee/Valet slices. Loss slices (very small or negative) are hidden.",
     "profitBreakdown.otherIncomeBreakdown[]"),

    ("4000-4004 (sales + volume), 5000-5005, 5041, 5046-5050 (cost)  grouped by month",
     "Aggregate Chart: Monthly Fuel Volume + Sales + Margin (stacked)",
     "Dashboard · Trends",
     "Monthly Fuel Performance Chart",
     "Monthly chart showing stacked fuel volume, sales and margin over the selected window.",
     "monthlyTrends[]"),

    ("Shop: 4032/4034/4036/4037/4039, 5016, 5032-5037, 5039, 5042  |  Valet: 4017, 4028-4031, 5015, 5028-5031, 5043, 5044  (grouped by month)",
     "Aggregate Chart: Monthly Shop & Valet Sales / Profit (signed) / Margin",
     "Dashboard · Shop & Valet Pillars",
     "Shop & Valet Monthly Combo Charts",
     "Monthly bar + line chart showing Shop and Coffee/Valet sales, profit and margin %. Profit values are signed (May 2026 fix) so loss months render as below-zero bars.",
     "monthlyTrends[].shopProfit · monthlyTrends[].valetProfit"),

    ("4100, 4101, 4102  (grouped by month)",
     "Aggregate Chart: Monthly Bunkering Totals",
     "Dashboard · Bunkering",
     "Bunkered Sales Chart",
     "Monthly bunkering totals — combines BP commission (4100), bunkered volume (4101) and bunkered commission (4102).",
     "bunkeredBreakdown[]"),

    ("UI-config only (no Sage ledger source)",
     "Configured: Marketing / Promotional Campaigns by Site",
     "Dashboard · Marketing",
     "Marketing Initiatives Table",
     "List of marketing / promotional campaigns by site. Configured in app settings, not pulled from the accounting ledger.",
     "marketingInitiatives[]"),

    ("4000-4004, 5000-5005, 5041, 5046-5050, 7150, 7151, 7200, 7800, 7906  (per month per site)",
     "Derived Chart: Monthly avgPPL + actualPPL Trend",
     "Dashboard · Forecast & Comparison",
     "PPL Comparison Chart (Monthly)",
     "Monthly chart comparing Avg PPL versus Actual PPL (after overheads) trend across selected sites.",
     "pplComparison[]"),

    # ============== COMPARISON PAGES ==============
    ("(All Gross Profit codes) + 7000-7005 (Labour) per dept_number",
     "Aggregate per-site: GP + Net Sales + PPL + Labour Cost (Site A vs Site B)",
     "Site Comparison Page",
     "Site-vs-Site Comparison Cards · Comparison Bar Chart · Pie Charts",
     "Side-by-side comparison of any two sites: Sales, Gross Profit, Volume, PPL, Labour Cost and margin %.",
     "site1Data · site2Data · profit · grossMarginPct"),

    ("(All Net Profit codes) + (All Misc Income codes) + (All Overheads codes) per dept_number  |  Excludes dept 0 (HO)",
     "Aggregate per-site: Total Net Profit + Sales + Volume + PPL after O/H (14-sites grid)",
     "Metrics Comparison Page",
     "14 Sites Grid - Sales / Gross Profit / Sale Volume / PPL after O/H",
     "Grid of all 14 active sites ranked by Sales, Gross Profit, Volume and PPL after Overheads. Head Office (dept 0) is excluded.",
     "sites[].profit · sites[].netSales · sites[].totalFuelVolume · sites[].pplAfterOverheads"),

    ("(Same codes as Metrics Comparison rows above)",
     "UI Control: Toggle visual ranking (bar) vs tabular grid (table)",
     "Metrics Comparison Page",
     "Chart View Toggle (Bar / Table)",
     "Switch between a visual chart view and a tabular grid view of the same site-ranking data.",
     "viewMode · sites[]"),

    # ============== AUTH / IDENTITY (non-financial) ==============
    ("hsrl_dashboard_users (id, email, password_hash, role, email_verified_at)",
     "Identity Ledger: Dashboard Users + Admin Users",
     "Auth / Identity",
     "Login (User) · Login (Admin) · Manage Users · Settings",
     "Dashboard user accounts. Passwords are encrypted before storage. Stored separately from accounting data.",
     "auth.user · auth.admin · /api/auth/login · /api/admin/users"),
]


# ----------------------------------------------------------------------------
# Workbook construction
# ----------------------------------------------------------------------------
def build_dataframe() -> pd.DataFrame:
    return pd.DataFrame(ROWS, columns=COLUMNS)


def autofit_column_widths(ws, df: pd.DataFrame, *, max_width: int = 80, padding: int = 4) -> None:
    """Set each column width based on the longest content (header or cell)."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].astype(str).tolist()),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + padding, max_width)


def apply_styles(ws, df: pd.DataFrame) -> None:
    """Deep Navy header, gridlines, alignment, wrap for long text."""
    NAVY = "1F3864"
    WHITE = "FFFFFF"

    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Calibri", size=10)
    mono_font = Font(name="Consolas", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_rows, n_cols = df.shape

    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.font = body_font
                cell.alignment = center
            elif col_idx == 6:
                cell.font = mono_font
                cell.alignment = left_wrap
            else:
                cell.font = body_font
                cell.alignment = left_wrap

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"


def main() -> None:
    df = build_dataframe()

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    apply_styles(ws, df)
    autofit_column_widths(ws, df)
    ws.sheet_view.showGridLines = True

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    print(f"OK  Wrote {len(df)} rows -> {OUTPUT_PATH}")
    print(f"    Columns: {', '.join(df.columns)}")


if __name__ == "__main__":
    main()
