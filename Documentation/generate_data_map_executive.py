"""
HSRL Dashboard — Executive Data Map Generator (Plain English Version)
=====================================================================

Generates a non-technical companion to HSRL_Dashboard_Accounting_Data_Map_Master.xlsx.
Same 6 columns, same 99 rows, same row numbering — but every cell is rewritten
in plain business English suitable for executive, finance and audit review.

Output file:
    HSRL_Dashboard_Data_Map_Executive_Summary.xlsx
    (written next to this script)

Usage:
    python Documentation/generate_data_map_executive.py
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "HSRL_Dashboard_Data_Map_Executive_Summary.xlsx"
SHEET_NAME = "Executive Data Map"

COLUMNS = [
    "Internal Accounting GL/Nominal Code",
    "Company Account Ledger Name",
    "HSRL Dashboard Module / Pillar",
    "Live UI Component / Chart Widget Name",
    "Data Extraction & Functional Transformation Logic",
    "Production API Frontend Payload Key",
]

# ----------------------------------------------------------------------------
# 99 rows — same row order as the technical workbook, plain-English content.
# (code, ledger_name, module, component, logic, where_shown)
# ----------------------------------------------------------------------------
ROWS = [
    # ============== FUEL SALES (5 grades) ==============
    ("4000", "Unleaded Petrol Sales", "Fuel Sales Section",
     "Total Sales figure on Home Screen, plus the Unleaded slice on the Fuel Mix chart",
     "All money taken from Unleaded petrol sales during the chosen dates. The litres sold come from the notes written on each sale entry.",
     "Shows on the Total Sales card and the Fuel Volume card."),

    ("4001", "Diesel Sales", "Fuel Sales Section",
     "Total Sales figure on Home Screen, Diesel slice on the Fuel Mix chart, Bunkered Sales chart",
     "All money taken from Diesel sales during the chosen dates. Bunkered Diesel litres come from a separate code (4101) and are added only to the volume number, not the money number.",
     "Shows on the Total Sales card and the Fuel Volume card."),

    ("4002", "Super Unleaded Sales", "Fuel Sales Section",
     "Total Sales figure on Home Screen, Super Unleaded slice on the Fuel Mix chart",
     "All money taken from Super Unleaded sales during the chosen dates. Litres come from the notes on each entry.",
     "Shows on the Total Sales card and the Fuel Volume card."),

    ("4003", "Super Diesel Sales", "Fuel Sales Section",
     "Total Sales figure on Home Screen, Super Diesel slice on the Fuel Mix chart",
     "All money taken from Super Diesel sales during the chosen dates. Litres come from the notes on each entry.",
     "Shows on the Total Sales card and the Fuel Volume card."),

    ("4004", "Adblue Sales", "Fuel Sales Section",
     "Total Sales figure on Home Screen, Adblue slice on the Fuel Mix chart",
     "All money taken from Adblue sales during the chosen dates. Litres come from the notes on each entry.",
     "Shows on the Total Sales card and the Fuel Volume card."),

    # ---- Bunkering ----
    ("4100", "BP Bunkering Commission", "Fuel Sales Section",
     "Bunkered Sales chart, Total Site Revenue card",
     "Commission earned from BP bunkering activity. Counted in the Total Site Revenue figure but NOT in the headline Fuel Sales money.",
     "Shows on the Bunkered Sales chart and the Total Revenue card."),

    ("4101", "Bunkered Sales (Volume Only)", "Fuel Sales Section",
     "Fuel Volume card, Bunkered Sales chart",
     "A volume-only line for bunkered fuel. Any reversal entries are automatically deducted. Does not add to the Fuel Sales money.",
     "Adds to the Fuel Volume total shown on the dashboard."),

    ("4102", "Bunkered Commission Income", "Fuel Sales Section",
     "Bunkered Sales chart",
     "Commission income from bunkered fuel deals.",
     "Shows on the Bunkered Sales chart."),

    # ---- Fuel purchase ----
    ("5000", "Cost of Unleaded Petrol Purchased", "Fuel Profit Section",
     "Gross Profit card, Net Profit card, EBITDA card",
     "Money paid to fuel suppliers for Unleaded petrol stock. Taken away from Unleaded sales to work out the fuel margin.",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5001", "Cost of Diesel Purchased", "Fuel Profit Section",
     "Gross Profit card, Net Profit card",
     "Money paid to fuel suppliers for Diesel stock. Paired with Diesel sales (code 4001) to work out the diesel margin.",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5002", "Cost of Super Unleaded Purchased", "Fuel Profit Section",
     "Gross Profit card",
     "Money paid for Super Unleaded stock. Paired with Super Unleaded sales (4002).",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5003", "Cost of Super Diesel Purchased", "Fuel Profit Section",
     "Gross Profit card",
     "Money paid for Super Diesel stock. Paired with Super Diesel sales (4003).",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5004", "Cost of Adblue Purchased", "Fuel Profit Section",
     "Gross Profit card",
     "Money paid for Adblue stock. Paired with Adblue sales (4004).",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5005", "Fuel Promotional Discounts", "Fuel Profit Section",
     "Gross Profit card, Net Profit card",
     "Promotional discounts given on fuel — for example loyalty programmes or pump-price reductions. Reduces fuel profit.",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5041", "Fuel Commission Paid", "Fuel Profit Section",
     "Gross Profit card, Net Profit card",
     "Fuel-related commission payments made out. Reduces fuel profit.",
     "Reduces the Fuel Profit figure shown on the dashboard."),

    ("5046", "Unleaded Stock Adjustment", "Fuel Profit Section",
     "Gross Profit card, ROI card",
     "Stock movement / adjustment line for Unleaded. Please confirm the exact name with your chart of accounts.",
     "Adjusts the Fuel Profit figure shown on the dashboard."),

    ("5047", "Diesel Stock Adjustment", "Fuel Profit Section",
     "Gross Profit card, ROI card",
     "Stock movement / adjustment line for Diesel. Please confirm the exact name with your chart of accounts.",
     "Adjusts the Fuel Profit figure shown on the dashboard."),

    ("5048", "Super Unleaded Stock Adjustment", "Fuel Profit Section",
     "Gross Profit card, ROI card",
     "Stock movement / adjustment line for Super Unleaded. Please confirm the exact name with your chart of accounts.",
     "Adjusts the Fuel Profit figure shown on the dashboard."),

    ("5049", "Super Diesel Stock Adjustment", "Fuel Profit Section",
     "Gross Profit card, ROI card",
     "Stock movement / adjustment line for Super Diesel. Please confirm the exact name with your chart of accounts.",
     "Adjusts the Fuel Profit figure shown on the dashboard."),

    ("5050", "General Stock Movement", "Fuel Profit Section",
     "Gross Profit card, ROI card",
     "General stock movement line for fuel.",
     "Adjusts the Fuel Profit figure shown on the dashboard."),

    # ============== SHOP SALES ==============
    ("4032", "E-Pay Sales Income", "Shop Sales Section",
     "Shop Profit card, Gross Profit Breakdown popup",
     "Money taken from E-Pay sales (mobile top-ups and utility payments).",
     "Shows on the Shop Sales card and Gross Profit breakdown."),

    ("4034", "Paypoint / Keycharge Sales Income", "Shop Sales Section",
     "Shop Profit card",
     "Money taken from Paypoint and Keycharge sales.",
     "Shows on the Shop Sales card."),

    ("4036", "Online Lottery Sales", "Shop Sales Section",
     "Shop Profit card",
     "Money taken from online lottery sales.",
     "Shows on the Shop Sales card."),

    ("4037", "Instant Lottery / Scratchcard Sales", "Shop Sales Section",
     "Shop Profit card",
     "Money taken from instant lottery and scratchcard sales.",
     "Shows on the Shop Sales card."),

    ("4039", "EV Charging Revenue", "Shop Sales Section",
     "Shop Profit card",
     "Money received from EV charging usage. Paired with EV running cost (code 5039).",
     "Shows on the Shop Sales card."),

    # ============== SHOP COSTS ==============
    ("5016", "Grocery Purchase Cost", "Shop Profit Section",
     "Shop Profit card, Gross Profit Breakdown popup",
     "Cost of grocery stock purchased. Added to the Shop Costs total in May 2026 — was missing before. Reduces Shop Profit.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5032", "E-Pay Purchase Cost", "Shop Profit Section",
     "Shop Profit card, Gross Profit Breakdown popup",
     "Cost of E-Pay stock purchased. Paired with E-Pay sales (code 4032).",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5033", "E-Pay Commission Paid", "Shop Profit Section",
     "Shop Profit card",
     "Commission paid out on E-Pay transactions.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5034", "Paypoint / Keycharge Purchase Cost", "Shop Profit Section",
     "Shop Profit card",
     "Cost of Paypoint / Keycharge stock. Paired with Paypoint sales (code 4034).",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5035", "Paypoint / Keycharge Commission Paid", "Shop Profit Section",
     "Shop Profit card",
     "Commission paid out on Paypoint / Keycharge transactions.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5036", "Online Lottery Cost", "Shop Profit Section",
     "Shop Profit card",
     "Cost of online lottery sales. Paired with sales code 4036.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5037", "Instant Lottery Cost", "Shop Profit Section",
     "Shop Profit card",
     "Cost of instant lottery / scratchcards. Paired with sales code 4037.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5039", "EV Charging Running Cost", "Shop Profit Section",
     "Shop Profit card",
     "EV charging running costs (electricity, network fees). Paired with sales code 4039.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    ("5042", "Lottery Operator Commission", "Shop Profit Section",
     "Shop Profit card",
     "Combined commission paid to the National Lottery operator.",
     "Reduces the Shop Profit figure shown on the dashboard."),

    # ============== COFFEE & VALET — SALES ==============
    ("4017", "Hot Food and Costa Coffee Sales", "Coffee & Valet Section",
     "Valet Profit card, Gross Profit Breakdown popup",
     "Money taken from hot food and Costa Coffee sales. Paired with the cost code 5015.",
     "Shows on the Coffee & Valet card and Gross Profit breakdown."),

    ("4028", "Car Wash Income", "Coffee & Valet Section",
     "Valet Profit card",
     "Money taken from Car Wash sales. Paired with cost code 5028.",
     "Shows on the Coffee & Valet card."),

    ("4029", "Jet Wash Income", "Coffee & Valet Section",
     "Valet Profit card",
     "Money taken from Jet Wash sales. Paired with cost code 5029.",
     "Shows on the Coffee & Valet card."),

    ("4030", "Car Vacuum Income", "Coffee & Valet Section",
     "Valet Profit card",
     "Money taken from car vacuum usage. Paired with cost code 5030.",
     "Shows on the Coffee & Valet card."),

    ("4031", "Car Tyre Air Pump Income", "Coffee & Valet Section",
     "Valet Profit card",
     "Money taken from the car tyre air pump. Paired with cost code 5031.",
     "Shows on the Coffee & Valet card."),

    # ============== COFFEE & VALET — COSTS ==============
    ("5015", "Hot Food / Costa Stock Cost", "Coffee & Valet Section",
     "Valet Profit card",
     "Cost of hot food and Costa coffee stock. Paired with sales code 4017.",
     "Reduces the Coffee & Valet Profit figure."),

    ("5028", "Car Wash Running Cost", "Coffee & Valet Section",
     "Valet Profit card",
     "Car Wash running costs (chemicals, water, maintenance).",
     "Reduces the Coffee & Valet Profit figure."),

    ("5029", "Jet Wash Running Cost", "Coffee & Valet Section",
     "Valet Profit card",
     "Jet Wash running costs.",
     "Reduces the Coffee & Valet Profit figure."),

    ("5030", "Car Vacuum Running Cost", "Coffee & Valet Section",
     "Valet Profit card",
     "Car vacuum running costs (electricity, maintenance).",
     "Reduces the Coffee & Valet Profit figure."),

    ("5031", "Car Tyre Air Pump Cost", "Coffee & Valet Section",
     "Valet Profit card",
     "Car tyre air pump running costs.",
     "Reduces the Coffee & Valet Profit figure."),

    ("5043", "Valet Commission Paid", "Coffee & Valet Section",
     "Valet Profit card",
     "Commission paid to third parties for valet services.",
     "Reduces the Coffee & Valet Profit figure."),

    ("5044", "Coffee Commission Paid", "Coffee & Valet Section",
     "Valet Profit card",
     "Commission paid to Costa Coffee on coffee sales.",
     "Reduces the Coffee & Valet Profit figure."),

    # ============== MISC / OTHER INCOME (13 lines) ==============
    ("4400", "Marketing Services Income", "Other Income Section",
     "EBITDA card, Other Income breakdown popup",
     "Income from marketing services provided. Counts towards Other Income inside EBITDA.",
     "Adds to the Other (Misc) Income figure."),

    ("4401", "ATM Cash Machine Income", "Other Income Section",
     "EBITDA card",
     "Income received from on-site ATM operators.",
     "Adds to the Other (Misc) Income figure."),

    ("4402", "Supplier Rebates Received", "Other Income Section",
     "EBITDA card",
     "Rebates from suppliers received during the period.",
     "Adds to the Other (Misc) Income figure."),

    ("4404", "Commissions Received", "Other Income Section",
     "EBITDA card",
     "Commission income earned from third-party services.",
     "Adds to the Other (Misc) Income figure."),

    ("4405", "Insurance Claims & Compensations", "Other Income Section",
     "EBITDA card, Other Income breakdown popup",
     "Insurance payouts and compensation received. Added to Other Income in May 2026 — was missing before.",
     "Adds to the Other (Misc) Income figure."),

    ("4407", "Rental Income", "Other Income Section",
     "EBITDA card",
     "Rent received from property and forecourt sublets. One of the largest steady income lines (around £89,000-£98,000 per month combined across sites).",
     "Adds to the Other (Misc) Income figure."),

    ("4410", "Head Office Miscellaneous Income", "Other Income Section",
     "EBITDA card",
     "Miscellaneous head office income. Sometimes contains one-off bookings (for example a £44,000 entry appeared in January 2026).",
     "Adds to the Other (Misc) Income figure."),

    ("4412", "Astwick Costa Coffee Rent", "Other Income Section",
     "EBITDA card",
     "Rent received from Costa Coffee at the Astwick site.",
     "Adds to the Other (Misc) Income figure."),

    ("4413", "EV Charging Rent / Revenue Share", "Other Income Section",
     "EBITDA card",
     "Rent or revenue share received from EV charge-point operators.",
     "Adds to the Other (Misc) Income figure."),

    ("4415", "Bank Interest Income", "Other Income Section",
     "EBITDA card",
     "Interest earned on bank deposits. The single largest Other Income line, around £90,000-£94,000 per month.",
     "Adds to the Other (Misc) Income figure."),

    ("4416", "ByBox Parcel Locker Income", "Other Income Section",
     "EBITDA card",
     "Income from ByBox parcel locker installations on site.",
     "Adds to the Other (Misc) Income figure."),

    ("4417", "Amazon Locker Rent", "Other Income Section",
     "EBITDA card",
     "Rent received from Amazon Locker installations on site.",
     "Adds to the Other (Misc) Income figure."),

    ("4418", "Euro Car Parks Rebate", "Other Income Section",
     "EBITDA card",
     "Rebate income from Euro Car Parks for forecourt parking enforcement.",
     "Adds to the Other (Misc) Income figure."),

    # ============== LABOUR ==============
    ("7000", "Staff Gross Wages", "Labour Cost Section",
     "Labour Cost card, Labour % chart",
     "Gross wages paid to all site and support staff. Used in the Labour Percent figure (Labour cost as a share of fuel sales).",
     "Shows on the Labour Cost card."),

    ("7001", "Employer NI Contributions on Staff", "Labour Cost Section",
     "Labour Cost card",
     "Employer National Insurance contributions paid on staff wages.",
     "Shows on the Labour Cost card."),

    ("7002", "Directors' Salaries", "Labour Cost Section",
     "Labour Cost card",
     "Salaries paid to company directors.",
     "Shows on the Labour Cost card."),

    ("7003", "Employer NI Contributions on Directors", "Labour Cost Section",
     "Labour Cost card",
     "Employer NI paid on directors' salaries.",
     "Shows on the Labour Cost card."),

    ("7005", "Directors' Pensions", "Labour Cost Section",
     "Labour Cost card",
     "Pension contributions paid for directors.",
     "Shows on the Labour Cost card."),

    # ============== OPERATING OVERHEADS (5 lines) ==============
    ("7150", "Rent", "Operating Overheads Section",
     "EBITDA card, PPL after Overheads card, Overhead Breakdown popup",
     "Property rent paid. One of the five operating overheads taken off EBITDA.",
     "Reduces the EBITDA figure shown on the dashboard."),

    ("7151", "Business Rates", "Operating Overheads Section",
     "EBITDA card, PPL after Overheads card, Overhead Breakdown popup",
     "Business rates paid to councils. Operating overhead in EBITDA.",
     "Reduces the EBITDA figure shown on the dashboard."),

    ("7200", "Electricity, Gas, Heating", "Operating Overheads Section",
     "EBITDA card, PPL after Overheads card, Overhead Breakdown popup",
     "Electricity, gas and heating costs across all sites. Operating overhead in EBITDA.",
     "Reduces the EBITDA figure shown on the dashboard."),

    ("7800", "Property Repairs & Equipment Renewals", "Operating Overheads Section",
     "EBITDA card, PPL after Overheads card, Overhead Breakdown popup",
     "Property repairs and equipment renewals. Operating overhead in EBITDA.",
     "Reduces the EBITDA figure shown on the dashboard."),

    ("7906", "Credit Card Processing Fees", "Operating Overheads Section",
     "EBITDA card, PPL after Overheads card, Overhead Breakdown popup",
     "Credit and debit card processing fees. Operating overhead in EBITDA.",
     "Reduces the EBITDA figure shown on the dashboard."),

    # ============== DEPRECIATION (7 lines) ==============
    ("8200", "Depreciation — Motor Vehicles", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on motor vehicles. Taken off EBITDA to reach Total Net Profit.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("8201", "Depreciation — Leasehold Land & Buildings", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on leasehold land and buildings. Taken off EBITDA.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("8202", "Depreciation — Freehold Land & Buildings", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on freehold land and buildings. Taken off EBITDA.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("8203", "Depreciation — Plant & Machinery", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on plant and machinery (pumps, tanks, equipment). Taken off EBITDA.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("8204", "Depreciation — Fixtures & Fittings", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on fixtures and fittings. Taken off EBITDA.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("8206", "Depreciation — Other Assets", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on other assets. Taken off EBITDA. (Note: code 8205 is not used in this chart of accounts.)",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("8207", "Depreciation — Site Development & Improvement", "Depreciation Section",
     "Total Net Profit Breakdown popup",
     "Depreciation expense on site development and improvement works. Taken off EBITDA.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    # ============== LOAN INTEREST + CORP TAX ==============
    ("7750", "Loan Interest Paid", "Finance Costs Section",
     "Total Net Profit Breakdown popup",
     "Interest paid on bank loans. Taken off EBITDA. Kept separate from operating overheads so PPL after Overheads is not double-counted.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    ("9000", "Corporation Tax Charge", "Tax Section",
     "Total Net Profit Breakdown popup",
     "Corporation tax charge for the period. Final amount taken off in the calculation: Total Net Profit = EBITDA minus Depreciation minus Loan Interest minus Corporation Tax.",
     "Reduces the Total Net Profit figure shown on the dashboard."),

    # ============== KPI CARDS (aggregate metrics) ==============
    ("Fuel codes 4000-4005 + 5000-5005 + 5041 + 5046-5050; Shop codes 4032,4034,4036,4037,4039 + 5016 + 5032-5037 + 5039 + 5042; Coffee & Valet codes 4017 + 4028-4031 + 5015 + 5028-5031 + 5043 + 5044",
     "Total Gross Profit (Fuel + Shop + Coffee & Valet, with losses correctly subtracting)",
     "Top KPI Strip",
     "Gross Profit headline card on the dashboard",
     "The headline Gross Profit figure for the business. Calculated as Fuel Profit plus Shop Profit plus Coffee and Valet Profit. If Shop or Valet runs at a loss, the loss correctly reduces the total (fixed in May 2026).",
     "The big Gross Profit number shown at the top of the dashboard."),

    ("Revenue codes 4000-4004, 4100-4102 | Cost codes 5000-5005, 5041, 5046-5050",
     "Net Profit (Total Revenue minus Total Cost)",
     "Top KPI Strip",
     "Net Profit headline card on the dashboard",
     "Net Profit equals Total Revenue minus Total Cost across the 14 Net Profit codes.",
     "The big Net Profit number shown at the top of the dashboard."),

    ("All Gross Profit codes plus Other Income codes 4400-4418 minus Operating Overheads 7150, 7151, 7200, 7800, 7906",
     "EBITDA (Gross Profit + Other Income minus Operating Overheads)",
     "Quick Insights",
     "EBITDA card and EBITDA breakdown popup",
     "Earnings before interest, tax, depreciation and amortisation. Calculation: Gross Profit plus Other Income minus Operating Overheads. Other Income now uses 13 codes (4405 added in May 2026).",
     "The EBITDA card and its breakdown popup on the dashboard."),

    ("Fuel codes 4000-4004 (litres from notes) + 4101 (bunkered volume) | Cost codes 5000-5005, 5041, 5046-5050",
     "Gross Profit Per Litre — Fuel Profit divided by Litres Sold (in pence)",
     "Quick Insights",
     "Gross PPL card on the dashboard",
     "Profit made on each litre of fuel sold, shown in pence. Calculation: Fuel Profit divided by Litres Sold, multiplied by 100. If litres are missing, the calculation falls back to sales value.",
     "The Gross PPL (pence per litre) card on the dashboard."),

    ("Fuel codes 4000-4004 + 5000-5004 (for fuel profit) plus Operating Overheads 7150, 7151, 7200, 7800, 7906",
     "PPL after Overheads — Profit per litre after running costs",
     "Quick Insights",
     "PPL after Overheads card on the dashboard",
     "The true profit per litre AFTER deducting operating overheads (rent, rates, electricity, repairs, card fees). Shows what the business actually earns per litre once running costs are paid.",
     "The PPL after Overheads card on the dashboard."),

    ("All EBITDA codes minus Depreciation 8200-8207 minus Loan Interest 7750 minus Corporation Tax 9000 | Investment codes configured separately",
     "Return on Investment — Net Profit divided by Investment, shown as a percentage",
     "ROI Page",
     "ROI headline card and the ROI Monthly Trend chart",
     "Return on Investment shown as a percentage. Calculation: Total Net Profit divided by the Investment value, multiplied by 100.",
     "The ROI card and ROI trend chart."),

    ("Site number list (department codes) | Evidence: any fuel sale on codes 4000-4004",
     "Active Sites — count of stations trading in the chosen period",
     "Quick Insights",
     "Active Sites card",
     "How many petrol stations had any fuel sales in the chosen period. Head Office is excluded from this count.",
     "The Active Sites number on the dashboard."),

    ("Bank account codes 1200, 1201, 1202, 1203, 1204, 1211-1222, 1240, 1250, 1251",
     "Bank Balance — most recent balance across all company accounts",
     "Quick Insights",
     "Bank Balance card",
     "Most recent bank balance across all company accounts — the head office Lloyds account, each site's account, and the deposit accounts.",
     "The Bank Balance card on the dashboard."),

    ("Fuel sales codes 4000-4004 divided by the count of active sites",
     "Average Sale Per Site — Total Sales divided by Number of Active Sites",
     "Quick Insights",
     "Average Sale Per Site card",
     "Average sales value per active station. Calculated as Total Sales divided by the number of active sites in the chosen period.",
     "The Average Sale Per Site card on the dashboard."),

    # ============== CHARTS ==============
    ("All Gross Profit codes plus all Other Income codes plus all Overhead codes, grouped by month",
     "Monthly Performance Trend — Gross Profit, EBITDA, Volume over time",
     "Trends Section",
     "Monthly Performance chart",
     "Trend chart showing Gross Profit, EBITDA and Fuel Volume month by month across the chosen window.",
     "The Monthly Performance trend chart on the dashboard."),

    ("Fuel codes 4000-4004 and 5000-5005, grouped by day",
     "Daily Sales, Volume and Profit chart",
     "Trends Section",
     "Date-Wise Data chart",
     "Daily breakdown of sales, fuel volume and profit across the chosen date range.",
     "The Date-Wise chart on the dashboard."),

    ("Unleaded: 4000+5000+5046; Diesel: 4001+4101+5001+5041+5047+4100+4102+5005; Super Unleaded: 4002+5002+5048; Super Diesel: 4003+5003+5049; Adblue: 4004+5004+5050",
     "Fuel Grade Mix — each grade's share of total fuel profit (Unleaded, Diesel, Super Unleaded, Super Diesel, Adblue, plus Shop and Coffee & Valet)",
     "Fuel Section",
     "Fuel Grade Mix pie chart",
     "Pie chart showing each fuel grade's share of total fuel profit, plus Shop and Coffee/Valet slices. Loss slices that are very small or negative are hidden from the pie.",
     "The Fuel Grade Mix pie chart on the dashboard."),

    ("Fuel codes 4000-4004 (sales and volume) and 5000-5005 (cost), grouped by month",
     "Monthly Fuel Performance — volume, sales and margin by month",
     "Trends Section",
     "Monthly Fuel Performance chart",
     "Monthly chart showing stacked fuel volume, sales and margin over the chosen window.",
     "The Monthly Fuel Performance chart on the dashboard."),

    ("Shop codes 4032, 4034, 4036, 4037, 4039 + 5016 + 5032-5037 + 5039 + 5042; Valet codes 4017 + 4028-4031 + 5015 + 5028-5031 + 5043 + 5044, grouped by month",
     "Monthly Shop and Coffee/Valet performance — sales, profit (with losses) and margin",
     "Shop and Coffee & Valet Sections",
     "Shop & Valet Monthly Combo charts",
     "Monthly bar plus line chart showing Shop and Coffee/Valet sales, profit and margin percent. Profit is now signed (fixed in May 2026) so loss months show as bars below zero.",
     "The Shop and Coffee/Valet monthly combo charts on the dashboard."),

    ("Bunkering codes 4100, 4101, 4102, grouped by month",
     "Monthly Bunkering Totals — BP commission, bunkered volume and bunkered commission",
     "Bunkering Section",
     "Bunkered Sales chart",
     "Monthly bunkering totals. Combines BP commission (4100), bunkered volume (4101) and bunkered commission (4102).",
     "The Bunkered Sales chart on the dashboard."),

    ("Configured in app settings (not from accounting data)",
     "Marketing & Promotional Campaigns by site (configured list)",
     "Marketing Section",
     "Marketing Initiatives table",
     "List of marketing and promotional campaigns by site. Configured in app settings, not pulled from the accounting ledger.",
     "The Marketing Initiatives table on the dashboard."),

    ("Fuel codes 4000-4004 + 5000-5005 + 5041 + 5046-5050 plus Operating Overheads 7150, 7151, 7200, 7800, 7906 per month per site",
     "Monthly Gross PPL vs Actual PPL (after Overheads) trend",
     "Forecast & Comparison Section",
     "PPL Comparison chart",
     "Monthly chart comparing Gross PPL versus PPL after Overheads, trend across the selected sites.",
     "The PPL Comparison chart on the dashboard."),

    # ============== COMPARISON PAGES ==============
    ("All Gross Profit codes plus Labour codes 7000-7005, broken down by site number",
     "Site-vs-Site Comparison — Sales, Gross Profit, Volume, PPL and Labour for two chosen sites",
     "Site Comparison Page",
     "Site comparison cards, comparison bar chart, comparison pie charts",
     "Side-by-side comparison of any two sites: Sales, Gross Profit, Volume, PPL, Labour Cost and margin percent.",
     "The Site Comparison page (Site A versus Site B)."),

    ("All Net Profit codes plus all Other Income codes plus all Overhead codes, broken down by site number, excluding Head Office",
     "All-Sites Ranking — every site listed with Sales, Gross Profit, Volume and PPL after Overheads",
     "Metrics Comparison Page",
     "14 Sites Grid (Sales / Gross Profit / Volume / PPL after Overheads)",
     "Grid of all 14 active sites ranked by Sales, Gross Profit, Volume and PPL after Overheads. Head Office is excluded.",
     "The 14-sites grid on the Metrics Comparison page."),

    ("Same data as the rows above",
     "View Toggle — switch between visual chart and table",
     "Metrics Comparison Page",
     "Chart / Table View toggle",
     "Switch between a visual chart view and a tabular grid view of the same site-ranking data.",
     "The Chart/Table toggle on the Metrics Comparison page."),

    # ============== AUTH (non-financial) ==============
    ("Dashboard user accounts table (user id, email, encrypted password, role, verified status)",
     "Dashboard User Accounts — login records for end users and administrators",
     "Login & User Management",
     "Login screen (user), Admin Login screen, Manage Users page, Settings page",
     "Dashboard user accounts for end users and administrators. Passwords are encrypted before being stored. Held separately from the accounting data.",
     "The login screens and the Manage Users page."),
]


# ----------------------------------------------------------------------------
# Workbook construction (shared styling with the technical version)
# ----------------------------------------------------------------------------
def build_dataframe() -> pd.DataFrame:
    return pd.DataFrame(ROWS, columns=COLUMNS)


def autofit_column_widths(ws, df: pd.DataFrame, *, max_width: int = 80, padding: int = 4) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].astype(str).tolist()),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + padding, max_width)


def apply_styles(ws, df: pd.DataFrame) -> None:
    NAVY = "1F3864"
    WHITE = "FFFFFF"

    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_rows, n_cols = df.shape

    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font = body_font
            if col_idx == 1:
                cell.alignment = center
            else:
                cell.alignment = left_wrap

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"


def main() -> None:
    df = build_dataframe()
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    apply_styles(ws, df)
    autofit_column_widths(ws, df)
    ws.sheet_view.showGridLines = True

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    print(f"OK  Wrote {len(df)} rows -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
